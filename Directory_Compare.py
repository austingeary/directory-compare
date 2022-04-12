import os
from datetime import datetime
import numpy as np
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, numbers
import tkinter as tk
import time

#TODO: Launch from bat file

def standardUI():
    #User form window
    window = tk.Tk()
    window.title('Directory Comparison')
    window.geometry('700x400')
    window.configure(background='#edeeff')

    #User form labels
    l1 = tk.Label(window, text = "Product:").grid(row=0, column=0)
    l2 = tk.Label(window, text = "Module:").grid(row=1, column=0)
    l3 = tk.Label(window, text = "Base Release:").grid(row=2, column=0)
    l4 = tk.Label(window, text = "New Release:").grid(row=3, column=0)
    l5 = tk.Label(window, text = "PD Threshold:").grid(row=4, column=0)
    l6 = tk.Label(window, text = "File Extensions:").grid(row=5, column=0)
    status_label = tk.Label(window, text = "")

    #User form Entry boxes
    brdr = 5
    wdth = 90
    e1 = tk.Entry(window, border=brdr, width=wdth)
    e2 = tk.Entry(window, border=brdr, width=wdth)
    e3 = tk.Entry(window, border=brdr, width=wdth)
    e4 = tk.Entry(window, border=brdr, width=wdth)
    e5 = tk.Entry(window, border=brdr, width=wdth)
    e6 = tk.Entry(window, border=brdr, width=wdth)

    #User form default text
    e1.insert(0, r'FHB')
    e2.insert(0, r'MED')
    e3.insert(0, r'May20')
    e4.insert(0, r'Nov20')
    e5.insert(0, r'0.05')
    e6.insert(0, r'.dat,.csv,.txt,.xlsx,.mdb')

    def onClick():

        start_time = time.time()
        status_label['text'] = 'Running...'

        base_rel_date = datetime.strptime(e3.get(), "%b%y")
        new_rel_date = datetime.strptime(e4.get(), "%b%y")

        base_year = str(base_rel_date.year)
        new_year = str(new_rel_date.year)

        base_rel = e2.get() + '_' + base_rel_date.strftime("%b%y")
        new_rel = e2.get() + '_' + new_rel_date.strftime("%b%y")

        base_rel_prod = e1.get() + '_' + base_rel
        new_rel_prod = e1.get() + '_' + new_rel

        out_file_name = e2.get() + '_' + e3.get() + '-' + e4.get() + '_' + e1.get() + '_File_Compare.xlsx'

        base_dir = os.path.join('P:\HI\Product Creation\Products',base_year,base_rel,base_rel_prod)
        new_dir = os.path.join('P:\HI\Product Creation\Products',new_year,new_rel,new_rel_prod)
        pd_threshold = float(e5.get())
        file_exts = tuple(map(str, e6.get().split(',')))
        output_file = os.path.join('P:\HI\Product Creation\Products',new_year,new_rel,'Working Docs',out_file_name)

        print(output_file)

        base_root = os.path.basename(base_dir)
        new_root = os.path.basename(new_dir)

        #Initialize dictionary to hold all path/file information
        dict = {}

        #Walk base directory and find new counterparts
        dict = walkDir(base_dir, new_dir, dict, pd_threshold, file_exts)

        #Walk new directory for files absent in base
        dict = walkDir(new_dir, base_dir, dict, pd_threshold, file_exts, True)

        #print(dict)
        df = createOutput(dict, base_root, new_root, output_file)

        end_time = time.time()
        duration = end_time - start_time
        status_label['text'] = f'Done. This took {duration:.2f} seconds'
        
        return None

    #Button
    button = tk.Button(window, text='Run Comparison', command=onClick)

    #Geometry manager
    pdx = 10
    pdy = 10
    e1.grid(row=0, column=1, padx=pdx, pady=pdy)
    e2.grid(row=1, column=1, padx=pdx, pady=pdy)
    e3.grid(row=2, column=1, padx=pdx, pady=pdy)
    e4.grid(row=3, column=1, padx=pdx, pady=pdy)
    e5.grid(row=4, column=1, padx=pdx, pady=pdy)
    e6.grid(row=5, column=1, padx=pdx, pady=pdy)
    button.grid(row=6, column=1, padx=pdx, pady=pdy)
    status_label.grid(row=7, column=1, padx=pdx, pady=pdy)

    window.mainloop()
    return None

def genericUI():
    #User form window
    window = tk.Tk()
    window.title('Directory Comparison')
    window.geometry('700x400')
    window.configure(background='#edeeff')

    #User form labels
    l1 = tk.Label(window, text = "Base Directory:").grid(row=0, column=0)
    l2 = tk.Label(window, text = "New Directory:").grid(row=1, column=0)
    l3 = tk.Label(window, text = "PD Threshold:").grid(row=2, column=0)
    l4 = tk.Label(window, text = "File Extensions:").grid(row=3, column=0)
    l5 = tk.Label(window, text = "Output File:").grid(row=4, column=0)
    status_label = tk.Label(window, text = "")

    #User form Entry boxes
    brdr = 5
    wdth = 90
    e1 = tk.Entry(window, border=brdr, width=wdth)
    e2 = tk.Entry(window, border=brdr, width=wdth)
    e3 = tk.Entry(window, border=brdr, width=wdth)
    e4 = tk.Entry(window, border=brdr, width=wdth)
    e5 = tk.Entry(window, border=brdr, width=wdth)

    #User form default text
    e1.insert(0, r'P:\HI\Product Creation...')
    e2.insert(0, r'P:\HI\Product Creation...')
    e3.insert(0, r'0.05')
    e4.insert(0, r'.dat,.csv,.txt,.xlsx,.mdb')
    e5.insert(0, r'File_Compare.xlsx')

    def onClick():

        print('Start!')
        start_time = time.time()
        status_label['text'] = 'Running...'

        base_dir = os.path.normpath(e1.get())
        new_dir = os.path.normpath(e2.get())
        pd_threshold = float(e3.get())
        file_exts = tuple(map(str, e4.get().split(',')))
        output_file = os.path.normpath(e5.get())

        base_root = os.path.basename(base_dir)
        new_root = os.path.basename(new_dir)

        #Initialize dictionary to hold all path/file information
        dict = {}

        #Walk base directory and find new counterparts
        dict = walkDir(base_dir, new_dir, dict, pd_threshold, file_exts)

        print('Finished walking through base directories')
        #print(dict)

        #Walk new directory for files absent in base
        dict = walkDir(new_dir, base_dir, dict, pd_threshold, file_exts, True)

        print('Finished walking through new directories')
        #print(dict)

        df = createOutput(dict, base_root, new_root, output_file)

        print('Output File Created: {}'.format(output_file))

        end_time = time.time()
        duration = end_time - start_time
        status_label['text'] = f'Done. This took {duration:.2f} seconds'
        
        print('Done!')
        return None

    #Button
    button = tk.Button(window, text='Run Comparison', command=onClick)

    #Geometry manager
    pdx = 10
    pdy = 10
    e1.grid(row=0, column=1, padx=pdx, pady=pdy)
    e2.grid(row=1, column=1, padx=pdx, pady=pdy)
    e3.grid(row=2, column=1, padx=pdx, pady=pdy)
    e4.grid(row=3, column=1, padx=pdx, pady=pdy)
    e5.grid(row=4, column=1, padx=pdx, pady=pdy)
    button.grid(row=5, column=1, padx=pdx, pady=pdy)
    status_label.grid(row=6, column=1, padx=pdx, pady=pdy)

    window.mainloop()
    return None


#Walk all files in directory and find counterparts in other directory - then add info to dictionary
def walkDir(base_dir, new_dir, dict, pd_thresh, file_exts, reverse=False):
    for bpath, dirs, files in os.walk(base_dir):
        subpath = bpath[len(base_dir):]
        if subpath == '':
            subpath_name = 'root'
        else:
            subpath_name = subpath
        print('Working through path: {}'.format(subpath_name))
        if not(subpath_name in dict):
            dict[subpath_name] = {}
        for filename in files:
            bfile = bpath + os.sep + filename
            nfile = new_dir + subpath + os.sep + filename
            if bfile.endswith(file_exts) and not(filename in dict[subpath_name]):
                print('Looking at file: {}'.format(filename))
                if reverse:
                    dict[subpath_name][filename] = compareFiles(nfile, bfile, pd_thresh)
                else:
                    dict[subpath_name][filename] = compareFiles(bfile, nfile, pd_thresh)
    return dict

def compareFiles(bfile, nfile, pd_thresh):

    #Initialize all values to None
    bcount, bsize, ncount, nsize, count_diff, size_diff, count_pd, size_pd = None, None, None, None, None, None, None, None

    #Initialize conditions to True
    base_exist, new_exist, pd_accept = 'True', 'True', 'True'

    #File types whose records can be counted via Python script
    countable_exts = (".dat",".csv",".txt")

    #If base file exists...
    if os.path.exists(bfile):
        #Count Records
        if bfile.endswith(countable_exts):
            bcount = getFileLineCount(bfile)
        elif bfile.endswith('.xlsx'):
            bcount = getXlRowCount(bfile)
        #Get size
        bsize = os.path.getsize(bfile) / 1024
    #And if base file does not exist, flag this
    else:
        base_exist = 'False'

    #If new file exists...
    if os.path.exists(nfile):
        #Count Records
        if nfile.endswith(countable_exts):
            ncount = getFileLineCount(nfile)
        elif nfile.endswith('.xlsx'):
            ncount = getXlRowCount(nfile)
        #Get size
        nsize = os.path.getsize(nfile) / 1024
    #If new file does not exist, flag this
    else:
        new_exist = 'False'

    #If base and new files exist...
    if base_exist == 'True' and new_exist == 'True':
        #Calculate difference and PD of sizes
        size_diff, size_pd = round(nsize - bsize, 2), round(propDiff(bsize, nsize), 3)
        #Calculate difference and PD of record counts
        if bfile.endswith(countable_exts) or bfile.endswith('xlsx'):
            count_diff, count_pd = ncount - bcount, propDiff(bcount, ncount)
            count_pd = round(count_pd, 3)
            #If count pd is beyond threshold, flag this.
            if abs(count_pd) > pd_thresh:
                pd_accept = 'False'
        else:
            #If size pd is beyond threshold, flag this.
            if abs(size_pd) > pd_thresh:
                pd_accept = 'False'

    #Rounding sizes after being used in calculation
    if os.path.exists(bfile):
        bsize = round(bsize, 2)
    if os.path.exists(nfile):
        nsize = round(nsize, 2)

    comp_dict = {'Base_Size':bsize,'New_Size':nsize,'Size_Diff':size_diff,'Size_PD':size_pd,
                    'Base_Count':bcount,'New_Count':ncount,'Count_Diff':count_diff,'Count_PD':count_pd,
                    'PD_Acceptable':pd_accept, 'Base_Exist':base_exist, 'New_Exist':new_exist}

    #print(comp_dict)

    return comp_dict

def getFileLineCount(path):
    f = open(path)                  
    lines = 0
    buf_size = 1024 * 1024
    read_f = f.read # loop optimization
    buf = read_f(buf_size)
    while buf:
        lines += buf.count('\n')
        buf = read_f(buf_size)
    f.close()
    return lines

def getXlRowCount(path):
    wb = xl.load_workbook(path)
    rows = 0
    for ws in wb.worksheets:
        print('Worksheet: {}'.format(ws.title))
        rows += ws.max_row
    return rows

def propDiff(a, b):
    if a + b == 0:
        return 0
    else:
        return round((b - a)/((b + a)/2), 3)

def createOutput(dict, base_root, new_root, output_file):

    col_names = ['Base_Size','New_Size','Size_Diff','Size_PD',
    'Base_Count','New_Count','Count_Diff','Count_PD',
    'PD_Acceptable', 'Base_Exist', 'New_Exist']

    idx = pd.MultiIndex(levels=[[],[]],
                        codes=[[],[]],
                        names=[u'Path', u'File'])

    df = pd.DataFrame(index=idx, columns=col_names)

    for subpath in dict:
        for file in dict[subpath]:
            df.loc[(subpath, file),:] = dict[subpath][file]

    df.rename(columns={'Base_Exist':base_root + ' Exist','New_Exist': new_root + ' Exist'}, inplace=True)

    df.to_excel(output_file, freeze_panes=(1,2))

    wb = xl.load_workbook(output_file)
    ws = wb.active

    #Auto-sizing columns
    for column_cells in ws.columns: 
        unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, column_cells)) 
        length = max(len(str(cell.value)) for cell in unmerged_cells) 
        ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.2

    #Highlighting exceptions and setting certain columns to integer format
    red_color = 'ff8080'
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type = "solid")
    for row in ws.iter_rows():
        for cell in row:
            #print(cell.value)
            if cell.value == 'False':
                cell.fill = red_fill
            if row != 1 and cell.column in [3,4,5,7,8,9]:
                cell.number_format = numbers.BUILTIN_FORMATS[3]

    #Save changes
    wb.save(output_file)

    return df

#genericUI()
standardUI()